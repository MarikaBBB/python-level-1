import openpyxl as o
import matplotlib.pyplot as plt

def create_excel_file(filename):
    # Create a new excel file and add in some data
    workbook = o.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    
    # Add headers
    headers = ["Months", "Sales"]
    sheet.append(headers)
    
    # Add sample data
    data = [
        ["January", 150],
        ["February", 200],
        ["March", 180],
        ["April", 220],
        ["May", 170],
        ["June", 210],
        ["July", 230],
        ["August", 250],
        ["September", 190],
        ["October", 220],
        ["November", 240],
        ["December", 260]
    ]

    for row in data:
        sheet.append(row)
    
    workbook.save(filename)
    print(f"The excel file {filename} was created successfully!")
    
def read_excel_file(filename):
    workbook = o.load_workbook(filename)
    sheet = workbook.active
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    return data
  
def visualise_data(data):
    # Extract months and sales data
    months = [row[0] for row in data[1:]]  
    sales = [row[1] for row in data[1:]]
    
    plt.figure(figsize=(10,6))
    plt.plot(months, sales, marker='o', linestyle='-', color='blue')
    plt.title("Monthly Sales Data")
    plt.xlabel("Month")
    plt.ylabel("Sales")
    plt.grid(True)
    plt.show()
  
def main():
    filename = 'sales_data.xlsx'
    
    # Create Excel File
    create_excel_file(filename)
    
    # Read data from excel file
    data = read_excel_file(filename)
    print("Data from excel file: ")
    for row in data:
        print(row)
    
    # Visualise the data
    visualise_data(data)
    
if __name__ == "__main__":
    main()
